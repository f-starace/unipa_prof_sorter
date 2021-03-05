from tqdm import tqdm
import os
from docx import Document
from pprint import pprint
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


# open file excel
wb = load_workbook(filename='presenze_cdd.xlsx')
ws = wb.active

red_fill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')
green_fill = PatternFill(start_color='49f000',
                   end_color='49f000',
                   fill_type='solid')


yellow_fill = PatternFill(start_color='fdd700',
                   end_color='fdd700',
                   fill_type='solid')

while True:
    show_tier = input('Vuoi specificare la fascia di appartenenza? [y/n]:    ')
    if show_tier == 'y':
        show_tier = True
    elif show_tier == 'n':
        show_tier = False
    else:
        continue

    add_colors = input('Vuoi aggiungere i colori alle singole celle? [y/n]:    ')
    if add_colors == 'y':
        add_colors = True
    elif add_colors == 'n':
        add_colors = False
    else:
        continue

    break



# creation of Professor class

class Professor():
    def __init__(self, row_number):
        self.name = ws[f'B{row_number}'].value
        self.surname = ws[f'A{row_number}'].value.capitalize()
        self.row_num = row_number
    
    def identifier(self):
        split_name = self.name.split()
        if len(split_name) > 1:

            first_letters = [name[0] for name in split_name]
            name_initial_str = '.'.join(first_letters)
        else:
            name_initial_str = self.name[0]

        return self.surname + ' ' + name_initial_str

    def fill_table(self, status, tier, column_name ):


        if status == 'present':
            cell_val = f'{tier}ª fascia: PRESENTE'  if show_tier else 'PRESENTE'
            if add_colors:
                ws[f"{column_name}{self.row_num}"].fill = green_fill

        elif status == 'absent':
            cell_val = f'{tier}ª fascia: ASSENTE' if show_tier else 'ASSENTE'
            if add_colors:
                ws[f"{column_name}{self.row_num}"].fill = red_fill


        elif status == 'justified':
            cell_val = f'{tier}ª fascia: ASSENTE GIUSTIFICATO' if show_tier else 'ASSENTE GIUSTIFICATO'
            if add_colors:
                ws[f"{column_name}{self.row_num}"].fill = yellow_fill


        # print(f'found match at {column_name}{self.row_num}')
        # print(cell_val)


        ws[f"{column_name}{self.row_num}"] = cell_val


    
    


# list of professor names:
professors = []

print('Retreiving professors personal info...')
for i in range(2,223):
    professors.append(Professor(i))





# list of dates:



# pick word document
word_docs_folder = os.path.join(os.getcwd(), 'word_docs' )
word_docs = os.listdir(word_docs_folder)

for doc in tqdm(word_docs):

    doc_path = os.path.join(word_docs_folder, doc)
    basename = os.path.basename(doc)
    print('reading doc: ' +  basename)


    # removing extension from filename
    date_str = os.path.splitext(basename)[0]

    if os.path.splitext(basename)[1] != '.docx':
        continue


    print('converting title to datetime obj')
    date_dt = datetime.strptime(date_str, '%d.%m.%Y')
    print(date_dt)

    # retriving column if same date 
    dates_raw = ws[1]


    for date in dates_raw:
        if date_dt == date.value:
            column_num = date.column
            column_name = get_column_letter(column_num) # get column
            
            break


    # read word document

    doc = Document(doc_path)
    paragraphs = doc.paragraphs
    first_tier_index = None
    second_tier_index = None
    for index, paragraph in enumerate(paragraphs):
        if paragraph.text.strip().lower() == 'professori di i fascia':
            first_tier_index = index

        elif paragraph.text.strip().lower() == 'professori di ii fascia':
            second_tier_index = index

        elif paragraph.text.strip().lower() == 'ricercatori':
            researcher_index = index
            break
    
    print('looping through "professori di prima fascia"...\n')

    for index in range(first_tier_index, second_tier_index):
        paragraph_text = paragraphs[index].text

        if paragraph_text.startswith('Presenti'):
            for professor in professors:
                if professor.identifier() in paragraph_text:
                    professor.fill_table('present',1, column_name )

        elif paragraph_text.startswith('Assenti:'):
            for professor in professors:
                if professor.identifier() in paragraph_text:
                    professor.fill_table('absent',1, column_name )

        elif paragraph_text.startswith('Assenti giustificati'):
            for professor in professors:
                if professor.identifier() in paragraph_text:
                    professor.fill_table('justified',1, column_name )

    
    print('looping through "professori di seconda fascia"...\n')

    for index in range(second_tier_index, researcher_index):

        paragraph_text = paragraphs[index].text
        
        if paragraph_text.startswith('Presenti'):
            for professor in professors:
                if professor.identifier() in paragraph_text:
                    professor.fill_table('present',2, column_name )

        elif paragraph_text.startswith('Assenti:'):
            for professor in professors:
                if professor.identifier() in paragraph_text:
                    professor.fill_table('absent',2, column_name )

        elif paragraph_text.startswith('Assenti giustificati'):
            for professor in professors:
                if professor.identifier() in paragraph_text:
                    professor.fill_table('justified',2, column_name )

wb.save('presenze_cdd.xlsx')

        











